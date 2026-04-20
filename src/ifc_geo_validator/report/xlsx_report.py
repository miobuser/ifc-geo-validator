"""Excel (.xlsx) report export.

Produces a multi-sheet workbook that civil engineers can open and
share without Python tooling. Uses openpyxl (declared under the
``bcf`` extra which already pulls xsdata/lxml). If openpyxl is not
installed, a clear ImportError with install hint is raised.

Sheets:
  1. Übersicht — one row per element with PASS/FAIL colour code
  2. Messwerte — every L1/L3 measurement per element
  3. Regelprüfung — flat list of all L4 rule check results
  4. Metadaten — validator version, timestamp, ruleset, CRS
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any


def export_xlsx(
    results: list[dict],
    output_path: str,
    *,
    ifc_filename: str = "",
    ruleset_name: str = "",
    coordinate_system: dict | None = None,
    project_name: str = "",
    author: str = "",
) -> str:
    """Write a multi-sheet .xlsx workbook to `output_path`.

    Args:
        results: per-element validation results (as returned by
            app.run_validation / cli).
        output_path: target .xlsx path.
        ifc_filename: source IFC filename (metadata only).
        ruleset_name: name of the applied ruleset.
        coordinate_system: dict from get_coordinate_system.
        project_name / author: optional metadata fields.

    Returns:
        The output path string on success.

    Raises:
        ImportError: if openpyxl is not installed.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ImportError(
            "openpyxl required for .xlsx export. "
            "Install with:  pip install openpyxl"
        ) from exc

    # ── Styling primitives ───────────────────────────────────────
    HEADER_FILL = PatternFill("solid", fgColor="CB0231")  # B+S red
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    PASS_FILL = PatternFill("solid", fgColor="C8E6C9")    # light green
    FAIL_FILL = PatternFill("solid", fgColor="FFCDD2")    # light red
    WARN_FILL = PatternFill("solid", fgColor="FFE0B2")    # light orange
    BORDER = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )

    def _header_row(ws, columns: list[str], row: int = 1) -> None:
        for col_idx, name in enumerate(columns, start=1):
            cell = ws.cell(row=row, column=col_idx, value=name)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = BORDER

    def _auto_width(ws, columns: list[str]) -> None:
        for col_idx, name in enumerate(columns, start=1):
            letter = get_column_letter(col_idx)
            # Start from header length, grow to longest value (cap 40)
            max_len = len(str(name))
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx,
                                    min_row=2, values_only=True):
                if row[0] is not None:
                    max_len = max(max_len, len(str(row[0])))
            ws.column_dimensions[letter].width = min(max_len + 2, 40)

    wb = Workbook()

    # ── Sheet 1: Übersicht ───────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Übersicht"
    overview_cols = ["ID", "Name", "Rolle", "Status", "Volumen (m³)",
                     "Oberfläche (m²)", "Wasserdicht", "Regeln PASS",
                     "Regeln FAIL", "Regeln SKIP"]
    _header_row(ws1, overview_cols)
    for row_idx, r in enumerate(results, start=2):
        l1 = r.get("level1", {})
        l2 = r.get("level2", {})
        l4 = r.get("level4", {})
        s = l4.get("summary", {}) if l4 else {}

        if l4:
            if s.get("errors", 0) > 0:
                status, fill = "FAIL", FAIL_FILL
            elif s.get("failed", 0) > 0:
                status, fill = "WARN", WARN_FILL
            else:
                status, fill = "PASS", PASS_FILL
        elif "error" in r:
            status, fill = "ERROR", FAIL_FILL
        else:
            status, fill = "—", None

        row_values = [
            r.get("element_id"),
            _safe_cell(r.get("element_name", "")),
            _safe_cell(l2.get("element_role", "")),
            status,
            round(l1.get("volume", 0), 3) if l1 else None,
            round(l1.get("total_area", 0), 3) if l1 else None,
            "Ja" if l1.get("is_watertight") else "Nein",
            s.get("passed") if l4 else None,
            s.get("failed") if l4 else None,
            s.get("skipped") if l4 else None,
        ]
        for col_idx, v in enumerate(row_values, start=1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=v)
            cell.border = BORDER
            if fill and col_idx == 4:  # status column
                cell.fill = fill
                cell.font = Font(bold=True)
    _auto_width(ws1, overview_cols)
    ws1.freeze_panes = "A2"

    # ── Sheet 2: Messwerte ───────────────────────────────────────
    ws2 = wb.create_sheet("Messwerte")
    measurement_cols = [
        "ID", "Name", "Rolle",
        "Kronenbreite (mm)", "Kronenneigung (%)",
        "Wandstärke min (mm)", "Wandhöhe (m)",
        "Anzug (n:1)", "Lotabweichung (°)",
        "Fundamentbreite (mm)", "Min. Krümmungsradius (m)",
        "Einbindetiefe (m)",
    ]
    _header_row(ws2, measurement_cols)
    for row_idx, r in enumerate(results, start=2):
        if "error" in r:
            continue
        l2 = r.get("level2", {})
        l3 = r.get("level3", {})
        l6_ctx = r.get("level6_context", {})
        row_values = [
            r.get("element_id"),
            _safe_cell(r.get("element_name", "")),
            _safe_cell(l2.get("element_role", "")),
            _fmt(l3.get("crown_width_mm"), 1),
            _fmt(l3.get("crown_slope_percent"), 2),
            _fmt(l3.get("min_wall_thickness_mm"), 1),
            _fmt(l3.get("wall_height_m"), 2),
            _fmt(l3.get("front_inclination_ratio"), 2),
            _fmt(l3.get("front_plumbness_deg"), 2),
            _fmt(l3.get("foundation_width_mm"), 1),
            _fmt(l3.get("min_radius_m"), 2),
            _fmt(l6_ctx.get("foundation_embedment_m"), 2),
        ]
        for col_idx, v in enumerate(row_values, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=v)
            cell.border = BORDER
    _auto_width(ws2, measurement_cols)
    ws2.freeze_panes = "A2"

    # ── Sheet 3: Regelprüfung ────────────────────────────────────
    ws3 = wb.create_sheet("Regelprüfung")
    check_cols = ["Element ID", "Element Name", "Regel-ID",
                  "Regel-Name", "Status", "Schweregrad",
                  "Ist", "Soll", "Referenz", "Meldung"]
    _header_row(ws3, check_cols)
    row_idx = 2
    for r in results:
        if "error" in r or "level4" not in r:
            continue
        for c in r["level4"].get("checks", []):
            status = c.get("status", "")
            fill = (PASS_FILL if status == "PASS"
                    else FAIL_FILL if status == "FAIL"
                    else None)
            row_values = [
                r.get("element_id"),
                _safe_cell(r.get("element_name", "")),
                _safe_cell(c.get("id", c.get("rule_id", ""))),
                _safe_cell(c.get("name", "")),
                status,
                _safe_cell(c.get("severity", "")),
                _fmt(c.get("actual"), 2),
                _safe_cell(str(c.get("expected", c.get("check_expr", "")))),
                _safe_cell(c.get("reference", "")),
                _safe_cell(c.get("message", "")),
            ]
            for col_idx, v in enumerate(row_values, start=1):
                cell = ws3.cell(row=row_idx, column=col_idx, value=v)
                cell.border = BORDER
                if fill and col_idx == 5:
                    cell.fill = fill
                    cell.font = Font(bold=True)
            row_idx += 1
    _auto_width(ws3, check_cols)
    ws3.freeze_panes = "A2"

    # ── Sheet 4: Metadaten ───────────────────────────────────────
    ws4 = wb.create_sheet("Metadaten")
    crs = coordinate_system or {}
    try:
        from ifc_geo_validator import get_version
        version = get_version()
    except Exception:
        version = "unknown"
    meta = [
        ("Prüftool", "ifc-geo-validator"),
        ("Version", version),
        ("Zeitstempel", datetime.now().isoformat(timespec="seconds")),
        ("IFC-Datei", ifc_filename),
        ("Ruleset", ruleset_name),
        ("Koordinatensystem",
         crs.get("name", "nicht deklariert") if crs.get("has_crs")
         else "nicht deklariert"),
        ("Vertikal-Datum", crs.get("vertical_datum", "")),
        ("Projekt", project_name),
        ("Prüfer/in", author),
        ("Anzahl Elemente", len(results)),
    ]
    for row_idx, (key, value) in enumerate(meta, start=1):
        ws4.cell(row=row_idx, column=1, value=key).font = Font(bold=True)
        ws4.cell(row=row_idx, column=2, value=str(value) if value else "—")
    ws4.column_dimensions["A"].width = 24
    ws4.column_dimensions["B"].width = 50

    wb.save(output_path)
    return output_path


def _fmt(v, digits: int = 2):
    """Format a number, passing None through."""
    if v is None:
        return None
    try:
        return round(float(v), digits)
    except (TypeError, ValueError):
        return v


def _safe_cell(v):
    """Neutralise Excel-formula-injection leader chars in a string cell.

    Delegates to ``cli._sanitize_csv_cell`` so the CSV and XLSX
    exporters share the same rule. Excel interprets `=`, `+`, `-`,
    `@`, `\\t`, `\\r` as formula leaders — an IFC element named
    ``=HYPERLINK("http://evil","click")`` would otherwise execute on
    the user's machine when they open the workbook.
    """
    from ifc_geo_validator.cli import _sanitize_csv_cell
    return _sanitize_csv_cell(v)
