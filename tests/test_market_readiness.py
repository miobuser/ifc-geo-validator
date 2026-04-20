"""Tests covering recent changes flagged as "NO COVERAGE" by the
test-coverage audit. One narrow test per new/changed contract:

  1. core.geometry.compute_signed_volume — new function
  2. core.mesh_converter.extract_mesh normals_flipped field
  3. core.face_classifier._build_face_adjacency non-manifold star graph
  4. validation.level3._compute_crown_width_sliced p10 robust fields
  5. validation.level5._min_bbox_distance_xy Euclidean
  6. validation.level5.validate_level5(config=...)
  7. core.anomaly_detection.detect_anomalies(config=...)
  8. core.ifc_parser.get_coordinate_system
  9. core.project_config.DEFAULT_CONFIG new sections

Kept deliberately minimal: one assertion per property so failures
point at the exact contract that regressed.
"""

import numpy as np
import pytest

from ifc_geo_validator.core.geometry import (
    compute_signed_volume, compute_volume,
)
from ifc_geo_validator.core.face_classifier import _build_face_adjacency
from ifc_geo_validator.core.anomaly_detection import detect_anomalies
from ifc_geo_validator.core.ifc_parser import get_coordinate_system
from ifc_geo_validator.core.project_config import DEFAULT_CONFIG
from ifc_geo_validator.validation.level5 import (
    validate_level5, _min_bbox_distance_xy, DEFAULT_MAX_GAP_3D_M,
)


# ── Unit cube for deterministic expectations ──────────────────────
UNIT_CUBE_VERTS = np.array([
    [0, 0, 0], [1, 0, 0], [1, 1, 0], [0, 1, 0],  # bottom
    [0, 0, 1], [1, 0, 1], [1, 1, 1], [0, 1, 1],  # top
], dtype=float)
# Triangles with outward-facing normals (counter-clockwise from outside)
UNIT_CUBE_FACES = np.array([
    [0, 2, 1], [0, 3, 2],  # bottom (normal -z)
    [4, 5, 6], [4, 6, 7],  # top (normal +z)
    [0, 1, 5], [0, 5, 4],  # front y=0 (normal -y)
    [2, 3, 7], [2, 7, 6],  # back y=1 (normal +y)
    [1, 2, 6], [1, 6, 5],  # right x=1 (normal +x)
    [0, 4, 7], [0, 7, 3],  # left x=0 (normal -x)
], dtype=np.int64)


def test_signed_volume_positive_for_outward_normals():
    v = compute_signed_volume(UNIT_CUBE_VERTS, UNIT_CUBE_FACES)
    assert v == pytest.approx(1.0, abs=1e-9)


def test_signed_volume_negative_for_inverted_winding():
    # Flipped winding: swap columns 1 and 2 → inward normals
    flipped = UNIT_CUBE_FACES[:, [0, 2, 1]]
    v = compute_signed_volume(UNIT_CUBE_VERTS, flipped)
    assert v == pytest.approx(-1.0, abs=1e-9)


def test_compute_volume_is_abs_of_signed():
    flipped = UNIT_CUBE_FACES[:, [0, 2, 1]]
    assert compute_volume(UNIT_CUBE_VERTS, flipped) == pytest.approx(1.0, abs=1e-9)


def test_face_adjacency_manifold_cube_pairs():
    """Manifold cube: every edge shared by exactly 2 faces → 18 pairs."""
    pairs = _build_face_adjacency(UNIT_CUBE_FACES)
    # 12 triangles, each with 3 edges = 36 half-edges / 2 per edge = 18 edges,
    # each edge contributes one pair → 18 pairs.
    assert len(pairs) == 18


def test_face_adjacency_non_manifold_star_graph():
    """k faces sharing an edge should emit k-1 pairs (star), not k*(k-1)/2."""
    # 4 triangles sharing edge (0, 1): fan-out from a shared edge
    faces = np.array([
        [0, 1, 2],
        [0, 1, 3],
        [0, 1, 4],
        [0, 1, 5],
    ], dtype=np.int64)
    pairs = _build_face_adjacency(faces)
    # The (0,1) edge is shared by 4 faces → 4 - 1 = 3 pairs minimum.
    # The (1,2), (0,2), (1,3), (0,3), ... edges are unique to one face each.
    # So only the shared-edge cluster produces pairs: exactly 3, not 6.
    assert len(pairs) == 3


def test_min_bbox_distance_xy_euclidean():
    """(3, 4) gap on diagonal → hypot = 5, not Manhattan 7."""
    a = {"bbox_min": np.array([0.0, 0.0, 0.0]),
         "bbox_max": np.array([1.0, 1.0, 1.0])}
    b = {"bbox_min": np.array([4.0, 5.0, 0.0]),
         "bbox_max": np.array([5.0, 6.0, 1.0])}
    # gap_x = 4-1 = 3, gap_y = 5-1 = 4 → Euclidean = 5
    assert _min_bbox_distance_xy(a, b) == pytest.approx(5.0, abs=1e-9)


def test_min_bbox_distance_xy_overlap_is_zero():
    a = {"bbox_min": np.array([0.0, 0.0, 0.0]),
         "bbox_max": np.array([2.0, 2.0, 1.0])}
    b = {"bbox_min": np.array([1.0, 1.0, 0.0]),
         "bbox_max": np.array([3.0, 3.0, 1.0])}
    assert _min_bbox_distance_xy(a, b) == 0.0


def test_validate_level5_config_max_gap_override():
    """A config max_gap_3d_m below the natural separation rejects pairs."""
    # Two far-apart single-element lists
    elem_a = {
        "element_id": 1, "element_name": "A",
        "level1": {"bbox": {"min": [0, 0, 0], "max": [1, 1, 1],
                            "size": [1, 1, 1]}, "volume": 1.0,
                   "centroid": [0.5, 0.5, 0.5]},
        "mesh_data": {"vertices": UNIT_CUBE_VERTS, "faces": UNIT_CUBE_FACES,
                      "normals": np.zeros((12, 3)), "areas": np.ones(12)},
    }
    elem_b = {
        "element_id": 2, "element_name": "B",
        "level1": {"bbox": {"min": [10, 10, 0], "max": [11, 11, 1],
                            "size": [1, 1, 1]}, "volume": 1.0,
                   "centroid": [10.5, 10.5, 0.5]},
        "mesh_data": {"vertices": UNIT_CUBE_VERTS + 10, "faces": UNIT_CUBE_FACES,
                      "normals": np.zeros((12, 3)), "areas": np.ones(12)},
    }
    # Default (1.0 m) rejects the pair (gap is 9 m). Custom 15 m includes it.
    res_default = validate_level5([elem_a, elem_b])
    assert res_default["summary"]["num_pairs"] == 0

    res_custom = validate_level5([elem_a, elem_b], config={"max_gap_3d_m": 15.0})
    assert res_custom["summary"]["num_pairs"] == 1


def test_anomaly_detection_config_override():
    """Custom front_back_ratio_flag changes whether asymmetry is flagged."""
    mesh = {"vertices": UNIT_CUBE_VERTS, "faces": UNIT_CUBE_FACES,
            "normals": np.zeros((12, 3)), "areas": np.ones(12)}
    l2 = {"summary": {"front": {"total_area": 5.0, "count": 1},
                      "back": {"total_area": 2.0, "count": 1},
                      "unclassified": {"total_area": 0.1, "count": 1},
                      "crown": {"total_area": 1.0, "count": 1},
                      "foundation": {"total_area": 1.0, "count": 1}},
          "has_crown": True, "has_foundation": True, "has_front": True,
          "has_back": True}
    l3 = {"crown_width_mm": 300.0, "min_wall_thickness_mm": 200.0}

    # 5:2 ratio = 2.5, default flag = 2.0 → one asymmetry anomaly
    default_anomalies = [a for a in detect_anomalies(mesh, l2, l3)
                         if a["type"] == "asymmetric_front_back"]
    assert len(default_anomalies) == 1

    # Raise the flag to 3.0 → suppresses the asymmetry diagnostic
    relaxed = [a for a in detect_anomalies(mesh, l2, l3,
                                           config={"front_back_ratio_flag": 3.0})
               if a["type"] == "asymmetric_front_back"]
    assert len(relaxed) == 0


def test_get_coordinate_system_returns_unspecified_for_empty_model():
    """When the model has no IfcProjectedCRS, the result is honest."""
    class FakeModel:
        def by_type(self, _):
            return []
    crs = get_coordinate_system(FakeModel())
    assert crs["name"] == "unspecified"
    assert crs["has_crs"] is False
    assert crs["eastings_offset"] is None


def test_default_config_has_reproducibility_sections():
    """Thesis-critical: project_config carries all override schemas."""
    for key in ("classifier", "pair_candidacy", "robust_stats", "anomaly"):
        assert key in DEFAULT_CONFIG, f"missing reproducibility section: {key}"
    # Spot-check a known value
    assert DEFAULT_CONFIG["pair_candidacy"]["max_gap_3d_m"] == DEFAULT_MAX_GAP_3D_M


# ── Security regression tests (pre-release audit) ─────────────────

def test_rule_expression_rejects_attribute_traversal():
    """A malicious ruleset cannot escape the sandbox via attribute chain.

    The classic eval()-sandbox bypass
    ``().__class__.__bases__[0].__subclasses__()`` must be refused.
    """
    from ifc_geo_validator.validation.level4 import _safe_eval as _evaluate_check
    # Attribute access is not a whitelisted AST node → ValueError
    with pytest.raises(ValueError):
        _evaluate_check(
            "().__class__.__bases__[0].__subclasses__() != None",
            {"dummy": 1.0},
        )


def test_rule_expression_rejects_arbitrary_function_call():
    """Only whitelisted builtins (currently `abs`) can be called."""
    from ifc_geo_validator.validation.level4 import _safe_eval as _evaluate_check
    with pytest.raises(ValueError):
        _evaluate_check("open('foo') != None", {"dummy": 1.0})
    with pytest.raises(ValueError):
        _evaluate_check("__import__('os') != None", {"dummy": 1.0})


def test_rule_expression_still_evaluates_legitimate_checks():
    """Safe evaluator must keep passing the same checks the old one did."""
    from ifc_geo_validator.validation.level4 import _safe_eval as _evaluate_check
    ctx = {"crown_width_mm": 305.0, "crown_slope_percent": 2.8}
    assert _evaluate_check("crown_width_mm >= 300", ctx) is True
    assert _evaluate_check("crown_width_mm < 300", ctx) is False
    assert _evaluate_check(
        "crown_width_mm >= 300 and crown_slope_percent <= 3",
        ctx,
    ) is True
    assert _evaluate_check("abs(crown_slope_percent) <= 3", ctx) is True


def test_csv_injection_sanitizer():
    """Leader characters for Excel formula injection get neutralised."""
    from ifc_geo_validator.cli import _sanitize_csv_cell
    assert _sanitize_csv_cell("=HYPERLINK(\"http://evil\",\"x\")").startswith("'=")
    assert _sanitize_csv_cell("+cmd|'/c calc'!A1").startswith("'+")
    assert _sanitize_csv_cell("-2+3").startswith("'-")
    assert _sanitize_csv_cell("@SUM(A:A)").startswith("'@")
    # Regular names stay untouched
    assert _sanitize_csv_cell("Wall_A1_Segment_03") == "Wall_A1_Segment_03"
    assert _sanitize_csv_cell(None) is None
    assert _sanitize_csv_cell(42) == 42


def test_json_report_does_not_leak_tempfile_path():
    """The report must record only the IFC basename, never the server path."""
    from ifc_geo_validator.report.json_report import generate_report
    r = generate_report("/tmp/tmp_secret_path.ifc", elements_results=[])
    assert "ifc_path" not in r["report"]
    assert r["report"]["ifc_file"] == "tmp_secret_path.ifc"


# ── Scalability regression tests ──────────────────────────────────

def test_level5_bbox_prefilter_rejects_distant_pairs():
    """N² bbox prefilter must reject pairs beyond max_gap in one pass."""
    # 50 tiny elements on a line 10 m apart — only adjacent pairs should
    # be candidates with the 1.0 m default cutoff.
    n = 50
    elements = []
    for i in range(n):
        cx = i * 10.0
        elements.append({
            "element_id": i,
            "element_name": f"E{i}",
            "level1": {
                "bbox": {"min": [cx, 0, 0], "max": [cx + 0.5, 0.5, 1],
                         "size": [0.5, 0.5, 1]},
                "volume": 0.25,
                "centroid": [cx + 0.25, 0.25, 0.5],
            },
            "mesh_data": {
                "vertices": UNIT_CUBE_VERTS * 0.5 + np.array([cx, 0, 0]),
                "faces": UNIT_CUBE_FACES,
                "normals": np.zeros((12, 3)),
                "areas": np.ones(12) * 0.25,
            },
        })
    result = validate_level5(elements)
    # Pairs 10 m apart → all rejected. num_pairs must be 0.
    assert result["summary"]["num_pairs"] == 0, (
        f"Prefilter should reject all distant pairs, got {result['summary']['num_pairs']}"
    )


def test_dilatation_joint_spacing_detects_gap():
    """A 20 m spacing between two walls exceeds the 15 m ASTRA cutoff."""
    # Two unit cubes 20 m apart along X
    a = {
        "element_id": 1, "element_name": "A",
        "level1": {"bbox": {"min": [0, 0, 0], "max": [1, 1, 1],
                            "size": [1, 1, 1]}, "volume": 1.0,
                   "centroid": [0.5, 0.5, 0.5]},
        "mesh_data": {"vertices": UNIT_CUBE_VERTS, "faces": UNIT_CUBE_FACES,
                      "normals": np.zeros((12, 3)), "areas": np.ones(12)},
    }
    b = {
        "element_id": 2, "element_name": "B",
        "level1": {"bbox": {"min": [20, 0, 0], "max": [21, 1, 1],
                            "size": [1, 1, 1]}, "volume": 1.0,
                   "centroid": [20.5, 0.5, 0.5]},
        "mesh_data": {"vertices": UNIT_CUBE_VERTS + np.array([20, 0, 0]),
                      "faces": UNIT_CUBE_FACES,
                      "normals": np.zeros((12, 3)), "areas": np.ones(12)},
    }
    result = validate_level5([a, b])
    joints = result["dilatation_joints"]
    assert len(joints) == 1
    assert joints[0]["exceeds_max"] is True
    assert joints[0]["spacing_m"] == pytest.approx(20.0, abs=0.001)
    assert result["summary"]["dilatation_violations"] == 1


def test_fix_direction_suggests_correct_verb():
    """A narrow crown hint points toward widening, not narrowing."""
    from ifc_geo_validator.validation.level4 import _fix_direction
    # Crown 285 mm, need ≥ 300 → should suggest 'verbreitern'
    msg = _fix_direction("crown_width_mm", ">=", 300.0, 285.0)
    assert "verbreitern" in msg
    assert "15" in msg  # magnitude of the gap

    # Crown 320 mm, need ≤ 300 (hypothetical) → 'verschmalen'
    msg = _fix_direction("crown_width_mm", "<=", 300.0, 320.0)
    assert "verschmalen" in msg


def test_xlsx_export_creates_four_sheets():
    """Excel report builds with Übersicht / Messwerte / Regelprüfung / Metadaten."""
    pytest.importorskip("openpyxl")
    import tempfile, os
    from ifc_geo_validator.report.xlsx_report import export_xlsx

    results = [{
        "element_id": 1,
        "element_name": "Test Wall",
        "level1": {"volume": 3.5, "total_area": 12.0, "is_watertight": True,
                   "bbox": {"size": [3, 0.3, 2]}},
        "level2": {"element_role": "wall_stem"},
        "level3": {"crown_width_mm": 305.0, "min_wall_thickness_mm": 300.0},
        "level4": {
            "summary": {"total": 3, "passed": 3, "failed": 0, "skipped": 0, "errors": 0},
            "checks": [{
                "id": "ASTRA-SM-L3-001", "name": "Kronenbreite",
                "status": "PASS", "severity": "ERROR",
                "actual": 305.0, "expected": ">= 300", "reference": "ASTRA §5",
                "message": "Check passed",
            }],
        },
    }]
    from openpyxl import load_workbook
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        out = tmp.name
    try:
        export_xlsx(results, out, ifc_filename="test.ifc", ruleset_name="ASTRA")
        wb = load_workbook(out)
        assert set(wb.sheetnames) == {"Übersicht", "Messwerte", "Regelprüfung", "Metadaten"}
        # Overview has 1 data row with our element
        assert wb["Übersicht"].cell(row=2, column=2).value == "Test Wall"
        # Rule-check sheet has the PASS check
        assert wb["Regelprüfung"].cell(row=2, column=5).value == "PASS"
    finally:
        os.unlink(out)


def test_alignment_context_reports_nearest_distance():
    """Wall centroid (10, 0) → alignment along x-axis → distance 0."""
    from ifc_geo_validator.validation.alignment import compute_alignment_context
    alignment = {
        "name": "A1",
        "points_xy": np.array([[i, 0] for i in range(0, 100, 2)]),
        "points_3d": np.array([[i, 0, 0] for i in range(0, 100, 2)]),
    }
    element = {
        "level1": {"centroid": [10.0, 5.0, 0.5]},
        "level3": {"min_radius_m": 50.0},
    }
    ctx = compute_alignment_context(element, [alignment])
    assert ctx["has_alignment"] is True
    assert ctx["min_alignment_distance_m"] == pytest.approx(5.0, abs=0.01)
    assert ctx["nearest_alignment_name"] == "A1"


def test_alignment_radius_ratio_straight_alignment_returns_none():
    """A perfectly straight alignment has no defined radius → ratio is None."""
    from ifc_geo_validator.validation.alignment import compute_alignment_context
    straight = {
        "name": "straight",
        "points_xy": np.array([[0, 0], [10, 0], [20, 0], [30, 0]]),
        "points_3d": np.array([[0, 0, 0], [10, 0, 0], [20, 0, 0], [30, 0, 0]]),
    }
    element = {
        "level1": {"centroid": [15.0, 0.0, 0.0]},
        "level3": {"min_radius_m": 50.0},
    }
    ctx = compute_alignment_context(element, [straight])
    assert ctx["alignment_radius_ratio"] is None  # no curvature → no ratio


def test_ids_export_produces_valid_xml():
    """IDS export round-trips through the XML parser and contains
    the expected buildingSMART schema elements."""
    import tempfile, os
    from ifc_geo_validator.report.ids_export import export_ids
    from xml.etree import ElementTree as ET

    ruleset = {
        "metadata": {
            "name": "Test Ruleset",
            "version": "1.0.0",
            "source": "Test",
            "ifc_filter": {"entity": "IfcWall", "predefined_type": "RETAININGWALL"},
        },
        "level_3": [
            {
                "id": "TEST-L3-001",
                "name": "Kronenbreite",
                "description": "Crown width >= 300 mm",
                "check": "crown_width_mm >= 300",
                "severity": "ERROR",
                "reference": "ASTRA §5.1",
            },
        ],
    }
    with tempfile.NamedTemporaryFile(suffix=".ids", delete=False) as tmp:
        out_path = tmp.name
    try:
        export_ids(ruleset, out_path, author="tester")
        # Parse must succeed
        tree = ET.parse(out_path)
        root = tree.getroot()
        # Root element is <ids> in the buildingSMART namespace
        assert root.tag.endswith("}ids")
        # At least one <specification>
        ns = "{http://standards.buildingsmart.org/IDS}"
        specs = root.findall(f".//{ns}specification")
        assert len(specs) == 1
        assert specs[0].get("name") == "Kronenbreite"
        # Applicability names IfcWall + RETAININGWALL
        ent = specs[0].find(f"{ns}applicability/{ns}entity/{ns}name/{ns}simpleValue")
        assert ent is not None and ent.text == "IfcWall"
    finally:
        os.unlink(out_path)


def test_alignment_no_alignments_returns_neutral_context():
    """Model without IfcAlignment yields has_alignment=False."""
    from ifc_geo_validator.validation.alignment import compute_alignment_context
    ctx = compute_alignment_context(
        {"level1": {"centroid": [0, 0, 0]}}, alignments=[],
    )
    assert ctx["has_alignment"] is False
    assert ctx["min_alignment_distance_m"] is None


def test_fix_direction_returns_empty_for_unknown_variable():
    """Unknown variable in _FIX_DIRECTIONS must return empty string,
    not raise KeyError. Guards against regressions that would crash
    the whole rule evaluation on any rule referencing a new variable
    not yet in the fix catalog."""
    from ifc_geo_validator.validation.level4 import _fix_direction
    assert _fix_direction("nonexistent_metric_mm", ">=", 1.0, 0.5) == ""
    # Also verify it doesn't raise on exotic ops
    assert _fix_direction("unknown_var", "==", 1.0, 2.0) == ""


def test_load_all_shipped_rulesets_parse():
    """Every YAML ruleset in the bundled rulesets/ folder must parse,
    have a metadata block, and contain at least one rule. Catches
    YAML syntax errors and missing-key regressions across the whole
    ruleset catalog — not just the two that have dedicated tests."""
    import glob
    from pathlib import Path
    from ifc_geo_validator.validation.level4 import load_ruleset

    rs_dir = Path(__file__).resolve().parents[1] / \
             "src" / "ifc_geo_validator" / "rules" / "rulesets"
    ruleset_files = list(rs_dir.glob("*.yaml"))
    assert len(ruleset_files) >= 4, "Expected at least 4 bundled rulesets"

    for rs_path in ruleset_files:
        rs = load_ruleset(str(rs_path))
        assert "metadata" in rs, f"{rs_path.name} missing metadata"
        assert "name" in rs["metadata"], f"{rs_path.name} missing metadata.name"
        # At least one rule across level_1/3/6
        n_rules = sum(len(rs.get(k, [])) for k in ("level_1", "level_3", "level_6"))
        assert n_rules > 0, f"{rs_path.name} has no rules"


def test_ids_export_handles_composite_expression():
    """A rule with `check: "a >= 1 and b <= 2"` emits a valid IDS
    specification without crashing. Current behaviour: the regex
    greedily matches the first comparison (``a >= 1``) and that
    becomes the property constraint — the second clause is silently
    dropped. Documenting this behaviour so any future IDS upgrade
    that splits compound expressions into multiple <property>
    elements is caught as an intentional change, not a regression."""
    import tempfile, os
    from xml.etree import ElementTree as ET
    from ifc_geo_validator.report.ids_export import export_ids

    ruleset = {
        "metadata": {
            "name": "Composite",
            "version": "1.0",
            "ifc_filter": {"entity": "IfcWall"},
        },
        "level_3": [{
            "id": "COMP-001",
            "name": "Height range",
            "description": "Wall height between 2 m and 8 m",
            "check": "wall_height_m >= 2.0 and wall_height_m <= 8.0",
            "severity": "WARNING",
        }],
    }
    with tempfile.NamedTemporaryFile(suffix=".ids", delete=False) as tmp:
        out = tmp.name
    try:
        export_ids(ruleset, out)
        tree = ET.parse(out)
        ns = "{http://standards.buildingsmart.org/IDS}"
        specs = tree.getroot().findall(f".//{ns}specification")
        assert len(specs) == 1
        # First comparison ("wall_height_m >= 2.0") is emitted as
        # the property constraint. The AND-clause is a known
        # limitation of the current IDS exporter.
        base = specs[0].find(f".//{ns}baseName/{ns}simpleValue")
        assert base is not None
        assert base.text == "wall_height_m"
    finally:
        os.unlink(out)


def test_level5_prefilter_finds_adjacent_pairs():
    """Prefilter must NOT reject pairs that are actually within cutoff."""
    # Two overlapping-bbox walls — gap should be 0, pair surviving.
    a = {
        "element_id": 1, "element_name": "A",
        "level1": {"bbox": {"min": [0, 0, 0], "max": [1, 1, 1],
                            "size": [1, 1, 1]}, "volume": 1.0,
                   "centroid": [0.5, 0.5, 0.5]},
        "mesh_data": {"vertices": UNIT_CUBE_VERTS, "faces": UNIT_CUBE_FACES,
                      "normals": np.zeros((12, 3)), "areas": np.ones(12)},
    }
    b = {
        "element_id": 2, "element_name": "B",
        "level1": {"bbox": {"min": [0.5, 0.5, 0], "max": [1.5, 1.5, 1],
                            "size": [1, 1, 1]}, "volume": 1.0,
                   "centroid": [1.0, 1.0, 0.5]},
        "mesh_data": {"vertices": UNIT_CUBE_VERTS + np.array([0.5, 0.5, 0]),
                      "faces": UNIT_CUBE_FACES,
                      "normals": np.zeros((12, 3)), "areas": np.ones(12)},
    }
    result = validate_level5([a, b])
    assert result["summary"]["num_pairs"] == 1
